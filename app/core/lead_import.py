"""
app/core/lead_import.py
-----------------------
Parsers that turn the various import sources (CSV / Excel / pasted text /
inbound email) into a uniform list of lead dicts. Every source converges on
`normalize_row`, so header handling and field coercion live in ONE place and the
API layer just validates + persists the result.

Nothing here touches the database — these are pure functions, which keeps them
trivial to unit-test.
"""
import csv
import io
import json
import re
from email.utils import parseaddr
from typing import Any, Dict, List

# Canonical lead field  ->  accepted header aliases (compared case-/space-insensitively).
# Only `email` is required downstream; everything else is optional.
_FIELD_ALIASES: Dict[str, set] = {
    "email": {"email", "email address", "e mail", "mail", "mail id", "emailid", "e mail id"},
    "first_name": {"first name", "firstname", "first", "given name", "fname"},
    "last_name": {"last name", "lastname", "last", "surname", "lname"},
    "company": {"company", "company name", "organization", "organisation", "org"},
    "job_title": {"job title", "title", "designation", "role", "jobtitle", "position"},
    "phone": {"phone", "phone number", "mobile", "contact", "contact number", "mobile number"},
    "budget": {"budget", "deal size", "deal value", "amount", "value", "budget usd"},
}

# Reverse lookup: normalized header token -> canonical field.
_ALIAS_TO_FIELD: Dict[str, str] = {
    alias: field for field, aliases in _FIELD_ALIASES.items() for alias in aliases
}


def _norm_header(h: str) -> str:
    """Lowercase, trim, and collapse separators so 'First_Name' == 'first name'."""
    return re.sub(r"[\s_\-.]+", " ", (h or "").strip().lower()).strip()


def normalize_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map an arbitrary row (whatever headers the source used) onto canonical lead
    fields. Unrecognized columns are ignored; blank values are dropped; budget
    is coerced to a float when present. Returns only the fields we understand.
    """
    out: Dict[str, Any] = {}
    for key, value in raw.items():
        field = _ALIAS_TO_FIELD.get(_norm_header(str(key)))
        if not field:
            continue
        if value is None:
            continue
        text = str(value).strip()
        if text == "":
            continue
        if field == "budget":
            # Tolerate "₹5,00,000", "$1200", "1200.0" etc. — keep digits and dot.
            cleaned = re.sub(r"[^0-9.]", "", text)
            if cleaned in ("", "."):
                continue  # non-numeric budget → treat as "not provided", not an error
            try:
                out["budget"] = float(cleaned)
            except ValueError:
                continue
        else:
            out[field] = text
    return out


def parse_csv(text: str) -> List[Dict[str, Any]]:
    """Parse CSV text (with a header row) into normalized lead dicts."""
    reader = csv.DictReader(io.StringIO(text))
    return [normalize_row(row) for row in reader]


def parse_json(text: str) -> List[Dict[str, Any]]:
    """
    Parse JSON text into normalized lead dicts. Accepts either a top-level list
    of objects, or a single object (wrapped into a one-element list).
    """
    data = json.loads(text)
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ValueError("JSON must be an object or an array of objects.")
    rows = []
    for item in data:
        if not isinstance(item, dict):
            raise ValueError("Each JSON item must be an object.")
        rows.append(normalize_row(item))
    return rows


def parse_xlsx(content: bytes) -> List[Dict[str, Any]]:
    """
    Parse the first sheet of an .xlsx workbook (first row = headers) into
    normalized lead dicts. Uses openpyxl in read-only mode for low memory use.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h) if h is not None else "" for h in header]
        result = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue  # skip fully blank rows
            raw = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            result.append(normalize_row(raw))
        return result
    finally:
        wb.close()


def parse_inbound_email(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Turn an inbound-email webhook payload (SendGrid / Mailgun Inbound Parse
    style) into a single normalized lead dict. The SENDER becomes the lead:
    their address is the email, their display name is split into first/last, and
    the subject is kept as a note-ish company hint when present.

    Accepts the common field spellings used by the major providers:
      from | sender | From         -> "Jane Doe <jane@corp.com>"
      subject | Subject
    """
    sender = (
        payload.get("from")
        or payload.get("sender")
        or payload.get("From")
        or ""
    )
    display_name, addr = parseaddr(str(sender))
    row: Dict[str, Any] = {}
    if addr:
        row["email"] = addr
    if display_name:
        parts = display_name.split()
        row["first_name"] = parts[0]
        if len(parts) > 1:
            row["last_name"] = " ".join(parts[1:])
    return normalize_row(row)
